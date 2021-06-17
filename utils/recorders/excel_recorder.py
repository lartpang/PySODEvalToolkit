# -*- coding: utf-8 -*-
# @Time    : 2021/1/3
# @Author  : Lart Pang
# @GitHub  : https://github.com/lartpang

import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Thanks:
# - Python_Openpyxl: https://www.cnblogs.com/programmer-tlh/p/10461353.html
# - Python之re模块: https://www.cnblogs.com/shenjianping/p/11647473.html
class _BaseExcelRecorder(object):
    def __init__(self, xlsx_path: str):
        """
        提供写xlsx文档功能的基础类。主要基于openpyxl实现了一层更方便的封装。

        :param xlsx_path: xlsx文档的路径。
        """
        self.xlsx_path = xlsx_path
        if not os.path.exists(self.xlsx_path):
            print("We have created a new excel file!!!")
            self._initial_xlsx()
        else:
            print("Excel file has existed!")

    def _initial_xlsx(self):
        Workbook().save(self.xlsx_path)

    def load_sheet(self, sheet_name: str):
        wb = load_workbook(self.xlsx_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name, index=0)
        sheet = wb[sheet_name]

        return wb, sheet

    def append_row(self, sheet: Worksheet, row_data):
        assert isinstance(row_data, (tuple, list))
        sheet.append(row_data)

    def insert_row(self, sheet: Worksheet, row_data, row_id, min_col=1, interval=0):
        assert isinstance(row_id, int) and isinstance(min_col, int) and row_id > 0 and min_col > 0
        assert isinstance(row_data, (tuple, list)), row_data

        num_elements = len(row_data)
        row_data = iter(row_data)
        for row in sheet.iter_rows(
            min_row=row_id,
            max_row=row_id,
            min_col=min_col,
            max_col=min_col + (interval + 1) * (num_elements - 1),
        ):
            for i, cell in enumerate(row):
                if i % (interval + 1) == 0:
                    sheet.cell(row=row_id, column=cell.column, value=next(row_data))

    @staticmethod
    def merge_region(sheet: Worksheet, min_row, max_row, min_col, max_col):
        assert max_row >= min_row > 0 and max_col >= min_col > 0

        merged_region = (
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        )
        sheet.merge_cells(merged_region)

    @staticmethod
    def get_col_id_with_row_id(sheet: Worksheet, col_name: str, row_id):
        """
        从指定行中寻找特定的列名，并返回对应的列序号
        """
        assert isinstance(row_id, int) and row_id > 0

        for cell in sheet[row_id]:
            if cell.value == col_name:
                return cell.column
        raise ValueError(f"In row {row_id}, there is not the column {col_name}!")

    def get_row_id_with_col_name(self, sheet: Worksheet, row_name: str, col_name: str):
        """
        从指定列名字的一列中寻找指定行，返回对应的row_id, col_id, is_new_row
        """
        is_new_row = True
        col_id = self.get_col_id_with_row_id(sheet=sheet, col_name=col_name, row_id=1)

        row_id = 0
        for cell in sheet[get_column_letter(col_id)]:
            row_id = cell.row
            if cell.value == row_name:
                return (row_id, col_id), not is_new_row
        return (row_id + 1, col_id), is_new_row

    @staticmethod
    def get_row_id_with_col_id(sheet: Worksheet, row_name: str, col_id: int):
        """
        从指定序号的一列中寻找指定行
        """
        assert isinstance(col_id, int) and col_id > 0

        is_new_row = True
        row_id = 0
        for cell in sheet[get_column_letter(col_id)]:
            row_id = cell.row
            if cell.value == row_name:
                return row_id, not is_new_row
        return row_id + 1, is_new_row

    @staticmethod
    def format_string_with_config(string: str, repalce_config: dict = None):
        assert repalce_config is not None
        if repalce_config.get("lower"):
            string = string.lower()
        elif repalce_config.get("upper"):
            string = string.upper()
        elif repalce_config.get("title"):
            string = string.title()

        sub_rule = repalce_config.get("replace")
        if sub_rule:
            string = re.sub(pattern=sub_rule[0], repl=sub_rule[1], string=string)
        return string


class MetricExcelRecorder(_BaseExcelRecorder):
    def __init__(
        self,
        xlsx_path: str,
        sheet_name: str = None,
        repalce_config=None,
        row_header=None,
        dataset_names=None,
        metric_names=None,
    ):
        """
        向xlsx文档写数据的类

        :param xlsx_path: 对应的xlsx文档路径
        :param sheet_name: 要写入数据对应的sheet名字
            默认为 `results`
        :param repalce_config: 用于替换对应数据字典的键的模式，会被用于re.sub来进行替换
            默认为 dict(lower=True, replace=(r"[_-]", ""))
        :param row_header: 用于指定表格工作表左上角的内容，这里默认为 `["methods", "num_data"]`
        :param dataset_names: 对应的数据集名称列表
            默认为rgb sod的数据集合 ["pascals", "ecssd", "hkuis", "dutste", "dutomron"]
        :param metric_names: 对应指标名称列表
            默认为 ["smeasure","wfmeasure","mae","adpfm","meanfm","maxfm","adpem","meanem","maxem"]
        """
        super().__init__(xlsx_path=xlsx_path)
        if sheet_name is None:
            sheet_name = "results"

        if repalce_config is None:
            self.repalce_config = dict(lower=True, replace=(r"[_-]", ""))
        else:
            self.repalce_config = repalce_config

        if row_header is None:
            row_header = ["methods", "num_data"]
        self.row_header = [
            self.format_string_with_config(s, self.repalce_config) for s in row_header
        ]
        if dataset_names is None:
            dataset_names = ["pascals", "ecssd", "hkuis", "dutste", "dutomron"]
        self.dataset_names = [
            self.format_string_with_config(s, self.repalce_config) for s in dataset_names
        ]
        if metric_names is None:
            metric_names = [
                "smeasure",
                "wfmeasure",
                "mae",
                "adpfm",
                "meanfm",
                "maxfm",
                "adpem",
                "meanem",
                "maxem",
            ]
        self.metric_names = [
            self.format_string_with_config(s, self.repalce_config) for s in metric_names
        ]

        self.sheet_name = sheet_name
        self._initial_table()

    def _initial_table(self):
        wb, sheet = self.load_sheet(sheet_name=self.sheet_name)

        # 插入row_header
        self.insert_row(sheet=sheet, row_data=self.row_header, row_id=1, min_col=1)
        # 合并row_header的单元格
        for col_id in range(len(self.row_header)):
            self.merge_region(
                sheet=sheet, min_row=1, max_row=2, min_col=col_id + 1, max_col=col_id + 1
            )

        # 插入数据集信息
        self.insert_row(
            sheet=sheet,
            row_data=self.dataset_names,
            row_id=1,
            min_col=len(self.row_header) + 1,
            interval=len(self.metric_names) - 1,
        )

        # 插入指标信息
        start_col = len(self.row_header) + 1
        for i in range(len(self.dataset_names)):
            self.insert_row(
                sheet=sheet,
                row_data=self.metric_names,
                row_id=2,
                min_col=start_col + i * len(self.metric_names),
            )
        wb.save(self.xlsx_path)

    def _format_row_data(self, row_data: dict) -> list:
        row_data = {
            self.format_string_with_config(k, self.repalce_config): v for k, v in row_data.items()
        }
        return [row_data[n] for n in self.metric_names]

    def __call__(self, row_data: dict, dataset_name: str, method_name: str):
        dataset_name = self.format_string_with_config(dataset_name, self.repalce_config)
        assert (
            dataset_name in self.dataset_names
        ), f"{dataset_name} is not contained in {self.dataset_names}"

        # 1 载入数据表
        wb, sheet = self.load_sheet(sheet_name=self.sheet_name)
        # 2 搜索method_name是否存在，如果存在则直接寻找对应的行列坐标，如果不存在则直接使用新行
        dataset_col_start_id = self.get_col_id_with_row_id(
            sheet=sheet, col_name=dataset_name, row_id=1
        )
        (method_row_id, method_col_id), is_new_row = self.get_row_id_with_col_name(
            sheet=sheet, row_name=method_name, col_name="methods"
        )
        # 3 插入方法名字到对应的位置
        if is_new_row:
            sheet.cell(row=method_row_id, column=method_col_id, value=method_name)
        # 4 格式化指标数据部分为合理的格式，并插入表中
        row_data = self._format_row_data(row_data=row_data)
        self.insert_row(
            sheet=sheet, row_data=row_data, row_id=method_row_id, min_col=dataset_col_start_id
        )
        # 4 写入新表
        wb.save(self.xlsx_path)
