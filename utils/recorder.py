# -*- coding: utf-8 -*-
# @Time    : 2020/7/8
# @Author  : Lart Pang
# @GitHub  : https://github.com/lartpang

# todo: xlsx_recoder
import os
from datetime import datetime

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook

from metrics.sod.metrics import MAE, Emeasure, Fmeasure, Smeasure, WeightedFmeasure


class MetricRecorder:
    def __init__(self):
        self.mae = MAE()
        self.fm = Fmeasure()
        self.sm = Smeasure()
        self.em = Emeasure()
        self.wfm = WeightedFmeasure()

    def update(self, pre: np.ndarray, gt: np.ndarray):
        assert pre.dtype == np.uint8, pre.dtype
        assert gt.dtype == np.uint8, gt.dtype

        self.mae.step(pre, gt)
        self.sm.step(pre, gt)
        self.fm.step(pre, gt)
        self.em.step(pre, gt)
        self.wfm.step(pre, gt)

    def show(self, bit_num=3) -> dict:
        fm_info = self.fm.get_results()
        fm = fm_info["fm"]
        pr = fm_info["pr"]
        wfm = self.wfm.get_results()["wfm"]
        sm = self.sm.get_results()["sm"]
        em = self.em.get_results()["em"]
        mae = self.mae.get_results()["mae"]
        results = {
            "em": em["curve"],
            "fm": fm["curve"],
            "p": pr["p"],
            "r": pr["r"],
            "Sm": sm,
            "wFm": wfm,
            "MAE": mae,
            "adpEm": em["adp"],
            "adpFm": fm["adp"],
        }
        if isinstance(bit_num, int):
            results = {k: v.round(bit_num) for k, v in results.items()}
        return results


class CurveDrawer:
    def __init__(self, row_num, col_num):
        self.fig = plt.figure()
        self.row_num = row_num
        self.col_num = col_num
        self.font_cfg = {
            "title": {
                "fontsize": 12,
                "fontweight": "bold",
                "fontname": "Liberation Sans",
            },
            "label": {
                "fontsize": 10,
                "fontweight": "normal",
                "fontname": "Liberation Sans",
            },
            "ticks": {
                "fontsize": 8,
                "fontweight": "normal",
                "fontname": "Liberation Sans",
            },
            "legend": {
                "size": 8,
                "weight": "normal",
                "family": "Liberation Sans",
            },
        }

    def add_subplot(self, curr_idx):
        assert isinstance(curr_idx, int) and curr_idx > 0
        self.ax = self.fig.add_subplot(self.row_num, self.col_num, curr_idx)
        self.ax.grid(True)

    def draw_method_curve(
        self,
        dataset_name,
        method_curve_setting,
        x_label,
        y_label,
        x_data,
        y_data,
        x_lim: tuple = (0, 1),
        y_lim: tuple = (0.1, 1),
    ):
        """
        :param method_curve_setting:  {
                "line_color": "color"(str),
                "line_style": "style"(str),
                "line_label": "label"(str),
                "line_width": width(int),
            }
        """
        # give plot a title
        self.ax.set_title(dataset_name, fontdict=self.font_cfg["title"])

        # make axis labels
        self.ax.set_xlabel(x_label, fontdict=self.font_cfg["label"])
        self.ax.set_ylabel(y_label, fontdict=self.font_cfg["label"])

        # 对坐标刻度的设置
        label = [f"{x:.2f}" for x in np.linspace(0, 1, 11)]
        self.ax.set_xticks(np.linspace(0, 1, 11))
        self.ax.set_yticks(np.linspace(0, 1, 11))
        self.ax.set_xticklabels(labels=label, fontdict=self.font_cfg["ticks"])
        self.ax.set_yticklabels(labels=label, fontdict=self.font_cfg["ticks"])

        self.ax.set_xlim(x_lim)
        self.ax.set_ylim(y_lim)

        # [CPFP, "red", "-", "OURS", 3],
        self.ax.plot(
            x_data,
            y_data,
            linewidth=method_curve_setting["line_width"],
            label=method_curve_setting["line_label"],
            color=method_curve_setting["line_color"],
            linestyle=method_curve_setting["line_style"],
        )

        # loc=0，自动将位置放在最合适的
        self.ax.legend(loc=3, prop=self.font_cfg["legend"])

        # 对坐标轴的框线进行设置, 设置轴
        self.ax.spines["top"].set_linewidth(1)
        self.ax.spines["bottom"].set_linewidth(1)
        self.ax.spines["left"].set_linewidth(1)
        self.ax.spines["right"].set_linewidth(1)

    def show(self):
        plt.show()


class TxtRecorder:
    def __init__(self, txt_path, resume=True):
        self.txt_path = txt_path
        mode = "a" if resume else "w"
        with open(txt_path, mode=mode, encoding="utf-8") as f:
            f.write(f" ========>> Date: {datetime.now()} <<======== \n")

    def add_row(self, row_name, row_data):
        with open(self.txt_path, mode="a", encoding="utf-8") as f:
            f.write(f" ========>> {row_name}: {row_data} <<======== \n")

    def add_method_results(
        self,
        data_dict: dict,
        method_name: str = "",
    ):
        # save the results under testing
        msg = method_name
        for k, v in data_dict.items():
            msg += f" {k} {v}\n"
        with open(self.txt_path, mode="a", encoding="utf-8") as f:
            f.write(msg + "\n")


class XLSXRecoder(object):
    def __init__(self, xlsx_path, data_mode="RGBD"):
        if data_mode == "RGB":
            self.dataset_list = [
                "DUTS",
                "DUT-OMRON",
                "HKU-IS",
                "ECSSD",
                "PASCAL-S",
                "SOC",
            ]
            self.dataset_num_list = [5019, 5168, 1447, 1000, 850, 1200]
        elif data_mode == "RGBD":
            self.dataset_list = [
                "lfsd",
                "njud",
                "nlpr",
                "rgbd135",
                "sip",
                "ssd",
                "stereo",
                "dutrgbd",
            ]
            self.dataset_num_list = [0, 0, 0, 0, 0, 0, 0, 0]
        else:
            raise NotImplementedError
        assert len(self.dataset_num_list) == len(self.dataset_list)

        self.metric_list = ["MaxF", "MeanF", "WFM", "MAE", "SM", "EM"]

        self.num_metrics = len(self.metric_list)
        self.num_datasets = len(self.dataset_list)

        self.path = xlsx_path
        if not os.path.exists(self.path):
            self.create_xlsx()
        else:
            print(
                f" !!! {xlsx_path} exists. Please confirm in advance that this file corresponds "
                f"to the "
                f"current data_mode({data_mode}), otherwise result data will not be written "
                f"normally."
            )

    def create_xlsx(self):
        # 创建一个Workbook对象
        wb = Workbook()
        # 创建一个Sheet对象
        sheet = wb.create_sheet(title="Results", index=0)
        # 获取活动的sheet
        sheet["A1"] = "name_dataset"
        sheet["A2"] = "num_dataset"

        for i, dataset_name in enumerate(self.dataset_list):
            if (i * self.num_metrics + 1) // 26 == 0:
                start_region_idx = f"{chr(ord('A') + (i * self.num_metrics + 1) % 26)}1"
            else:
                start_region_idx = (
                    f"{chr(ord('A') + (i * self.num_metrics + 1) // 26 - 1)}"
                    f"{chr(ord('A') + (i * self.num_metrics + 1) % 26)}1"
                )
            if ((i + 1) * self.num_metrics) // 26 == 0:
                end_region_idx = f"{chr(ord('A') + ((i + 1) * self.num_metrics) % 26)}1"
            else:
                end_region_idx = (
                    f"{chr(ord('A') + ((i + 1) * self.num_metrics) // 26 - 1)}"
                    f"{chr(ord('A') + ((i + 1) * self.num_metrics) % 26)}1"
                )
            region_idx = f"{start_region_idx}:{end_region_idx}"
            sheet.merge_cells(region_idx)  # 合并一行中的几个单元格
            sheet[start_region_idx] = dataset_name.upper()

            # 构造第二行数据
            start_region_idx = start_region_idx.replace("1", "2")
            sheet[start_region_idx] = self.dataset_num_list[i]

        # 构造第三行数据
        third_row = ["metrics"] + self.metric_list * self.num_datasets
        sheet.append(third_row)

        # 最后保存workbook
        wb.save(self.path)

    def write_xlsx(self, model_name, data):
        """
        向xlsx文件中写入数据
        :param model_name: 模型名字
        :param data: 数据信息，包含数据集名字和对应的测试结果
        """
        # 必须先得由前面的部分进行xlsx文件的创建，确保前三行OK满足要求，后面的操作都是从第四行开始的
        wb = load_workbook(self.path)
        assert "Results" in wb.sheetnames, (
            "Please make sure you are working with xlsx files created by " "`create_xlsx`"
        )
        sheet = wb["Results"]
        num_cols = self.num_metrics * self.num_datasets + 1

        if model_name in sheet["A"]:
            # 说明，该模型已经存在条目中，只需要更新对应的数据集结果即可
            idx_insert_row = sheet["A"].find(model_name)
        else:
            idx_insert_row = len(sheet["A"]) + 1
            sheet.cell(row=idx_insert_row, column=1, value=model_name)

        for dataset_name in data.keys():
            # 遍历每个单元格
            for row in sheet.iter_rows(min_row=1, min_col=2, max_col=num_cols, max_row=1):
                for cell in row:
                    if cell.value == dataset_name:
                        for i in range(self.num_metrics):
                            matric_name = sheet.cell(row=3, column=cell.column + i).value
                            sheet.cell(
                                row=idx_insert_row,
                                column=cell.column + i,
                                value=data[dataset_name][matric_name],
                            )
        wb.save(self.path)
